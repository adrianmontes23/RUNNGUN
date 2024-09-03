from tkinter import *
from winsound import *
from openpyxl import Workbook, load_workbook
import random
import time
import math
import bullets
import enemy

class RunNGun():
    
    def __init__(self):
        # Set up the root window, frame, and canvas
        self.tk = Tk()
        self.tk.title("RunN'Gun")
        self.tk.resizable(0, 0)
        # Create the frame and initialize the grid layout manager
        frame = Frame(self.tk)
        frame.grid()
        # Create a canvas
        self.canvas = Canvas(frame, width=1280, height=720)
        self.canvas.grid(row=1, column=1)
        self.wCanvas = Canvas(frame, width = 200, height = 720)
        self.wCanvas.grid(row = 1, column = 0)
        # Add the other widgets/set up the canvas
        self.dr = PhotoImage(file = "Other\¿̴̯̰̬̅̈́̈́́̉\d̸̞͉̅͝ȍ̵̝͙͔o̴͖̎ŗ̷͑͐̒.gif")
        self.crate = PhotoImage(file = "Other\¿̴̯̰̬̅̈́̈́́̉\C̷̜̻̈́r̸͙͛a̵̝̤̽̑ṱ̵̭̂e̴̛̥ͅ.gif")
        self.kDisp = PhotoImage(file = "wDisp\Knife.gif")
        self.pSR = PhotoImage(file = "pDisp\pSoldierR.gif")
        self.pSL = PhotoImage(file = "pDisp\pSoldierL.gif")
        self.uSR = PhotoImage(file = "pDisp\\uSoldierR.gif")
        self.uSL = PhotoImage(file = "pDisp\\uSoldierL.gif")
        self.shSR = PhotoImage(file = "pDisp\\shSoldierR.gif")
        self.shSL = PhotoImage(file = "pDisp\\shSoldierL.gif")
        self.arSR = PhotoImage(file = "pDisp\\arSoldierR.gif")
        self.arSL = PhotoImage(file = "pDisp\\arSoldierL.gif")
        self.snSR = PhotoImage(file = "pDisp\\snSoldierR.gif")
        self.snSL = PhotoImage(file = "pDisp\\snSoldierL.gif")
        self.dS = PhotoImage(file = "pDisp\deadSoldier.gif")
        self.kSR = PhotoImage(file = "pDisp\kSoldierR.gif")
        self.kSL = PhotoImage(file = "pDisp\kSoldierL.gif")
        self.bg = PhotoImage(file = "Other\\forestBG.gif")
        self.bP = PhotoImage(file = "Plataforms\Basic100x20.gif")
        self.dP = PhotoImage(file = "Plataforms\Dirt100x20.gif")
        self.h0 = PhotoImage(file = "Healths\helth0.gif")
        self.h1 = PhotoImage(file = "Healths\helth1.gif")
        self.h2 = PhotoImage(file = "Healths\helth2.gif")
        self.h3 = PhotoImage(file = "Healths\helth3.gif")
        self.AB3 = PhotoImage(file = "Healths\ABX3.gif")
        self.AB2 = PhotoImage(file = "Healths\ABX2.gif")
        self.AB1 = PhotoImage(file = "Healths\ABX1.gif")
        self.ABX = PhotoImage(file = "Healths\ABX.gif")
        self.lck = PhotoImage(file = "wDisp\lock.gif")
        self.ck = PhotoImage(file = "wDisp\Check.gif")
        self.looseDisp = PhotoImage(file = "Other\loose.gif")
        self.wbg = PhotoImage(file = "wDisp\weaponBG.gif")
        self.cbg = PhotoImage(file = "Other\creditScreen.gif")
        self.cfg = PhotoImage(file = "Other\creditThanks.gif")
        self.cT = PhotoImage(file = "Other\creditText.gif")
        self.gravitated = []
        self.colliders = []
        self.collidersI = []
        self.bulletL = []
        self.enemyL = []
        self.wCanvas.create_image(100, 360, image = self.wbg)
        self.check = self.wCanvas.create_image(100, 100, image = self.ck)
        self.uL = self.wCanvas.create_image(100, 230, image = self.lck)
        self.uT = self.wCanvas.create_text(100, 230, font = ("Times", "25"), text = "50 Coins", fill = "yellow")
        self.shL = self.wCanvas.create_image(100, 350, image = self.lck)
        self.shT = self.wCanvas.create_text(100, 350, font = ("Times", "25"), text = "100 Coins", fill = "yellow")
        self.arL = self.wCanvas.create_image(100, 460, image = self.lck)
        self.arT = self.wCanvas.create_text(100, 460, font = ("Times", "25"), text = "200 Coins", fill = "yellow")
        self.snL = self.wCanvas.create_image(100, 580, image = self.lck)
        self.snT = self.wCanvas.create_text(100, 580, font = ("Times", "25"), text = "350 Coins", fill = "yellow")
        self.canvas.create_image(640, 360, image = self.bg)
        self.placeMap()
        self.healthDisp = self.canvas.create_image(1180, 50, image = self.h3)
        self.abDisp = self.canvas.create_image(1180, 200, image = self.AB3)
        self.canvas.create_image(1180, 600, image = self.kDisp)
        self.coinDisp = self.canvas.create_text(20, 20, font = ("Times", "25"), anchor = "nw", text = "Coins: 0", fill = "Yellow")
        self.ammoDisp = self.canvas.create_text(60, 666, font = ("Times", "25"), text = "Pistol\n 7/7", fill = "Blue")
        self.rT = self.canvas.create_text(180, 666, font = ("Times", "25"), fill = "light Blue", text = "")
        self.gT = self.canvas.create_text(500, 500, font = ("Times", "25"), fill = "light Blue", text = "")
        self.canvas.bind_all("<KeyPress-w>", self.jump)
        self.canvas.bind_all("<KeyRelease-w>", self.jumpR)
        self.canvas.bind_all("<KeyPress-a>", self.a)
        self.canvas.bind_all("<KeyRelease-a>", self.aR)
        self.canvas.bind_all("<KeyPress-d>", self.d)
        self.canvas.bind_all("<KeyRelease-d>", self.dR)
        self.canvas.bind_all("<KeyPress-space>", self.shoot)
        self.canvas.bind_all("<KeyRelease-space>", self.shootR)
        self.canvas.bind_all("<KeyPress-1>", self.wSelect)
        self.canvas.bind_all("<KeyPress-2>", self.wSelect)
        self.canvas.bind_all("<KeyPress-3>", self.wSelect)
        self.canvas.bind_all("<KeyPress-4>", self.wSelect)
        self.canvas.bind_all("<KeyPress-5>", self.wSelect)
        self.canvas.bind_all("<KeyPress-f>", self.swingK)
        self.canvas.bind_all("<KeyPress-r>", self.heal)
        self.unlockedW = ["Pistol"]
        self.aDispT = "Pistol \n 7/7"
        self.direction = "r"
        self.weapon = "Pistol"
        self.ammoC = 7
        self.pAmmo = 7
        self.uAmmo = 32
        self.shAmmo = 8
        self.arAmmo = 30
        self.snAmmo = 5
        self.reloadT = 20
        self.slashing = False
        self.jumping = False
        self.jAble = False
        self.gameOver = False
        self.reloading = False
        self.gHand = False
        self.nadeTime = True
        self.gLive = False
        self.playerMove = False
        self.win = False
        self.shootA = 0
        self.coins = 5
        self.totalScore = 0
        self.gameOverR = 0
        self.healthA = 3
        self.ABA = 3
        self.nadeA = 3
        self.jTimer = 0
        self.gTimer = 50
        self.x = 0
        self.y = 0
        
        # Set up the stop button
        self.stopb = Button(frame, text="OFF", command=self.endProgram)
        self.stopb.grid(row=2, column=0, columnspan = 2)
        self.stop = False

    def placeMap(self):
        wb = Workbook()
        wb = load_workbook("Map.xlsx")
        ws = wb.active
        for c in range(1, ws.max_row + 1):
            char = chr(65)
            r = char + str(c)
            if ws[r].value == "Box":
                rn = ws[r].row
                for i in range(1, 5):
                    char = chr(65 + i)
                    coor = char + str(rn)
                    if i == 1:
                        a = ws[coor].value
                    if i == 2:
                        b = ws[coor].value
                    if i == 3:
                        c = ws[coor].value
                    if i == 4:
                        d = ws[coor].value
                if d-b == 20 and c-a == 20:
                    self.collidersI.append(self.canvas.create_image(a + 10, b + 10, image = self.crate))
                if d - b == 20:
                    if (c - a) % 100 == 0:
                        x = (c - a) / 100
                        xcord = 50
                        for i in range(int(x)):
                            self.collidersI.append(self.canvas.create_image(a+xcord, b+10, image = self.bP))
                            xcord += 100
                if (c - a) == 100:
                    if (d - b) % 20 == 0:
                        y = (d - b) / 20
                        ycord = 10
                        for i in range(int(y)):
                            if ycord == 10:
                                self.collidersI.append(self.canvas.create_image(a+50, b+ycord, image = self.bP))
                            else:
                                self.collidersI.append(self.canvas.create_image(a+50, b+ycord, image = self.dP))
                            ycord += 20
                if (d - b) % 20 == 0:
                    if (c - a) % 100 == 0:
                        y = (d - b) / 20
                        ycord = 10
                        for g in range(int(y)):
                            x = (c - a) / 100
                            xcord = 50
                            for i in range(int(x)):
                                if ycord == 10:
                                    self.collidersI.append(self.canvas.create_image(a+xcord, b+ycord, image = self.bP))
                                else:
                                    self.collidersI.append(self.canvas.create_image(a+xcord, b+ycord, image = self.dP))
                                xcord += 100
                            ycord += 20
                self.colliders.append(self.canvas.create_rectangle(a, b, c, d, width = 0))
            if ws[r].value == "Door":
                rn = ws[r].row
                for i in range(1,3):
                    char = chr(65 + i)
                    coor = char + str(rn)
                    if i == 1:
                        a = ws[coor].value
                    if i == 2:
                        b = ws[coor].value
                self.door = self.canvas.create_image(a, b, image = self.dr)
            if ws[r].value == "Player":
                rn = ws[r].row
                for i in range(1, 3):
                    char = chr(65 + i)
                    coor = char + str(rn)
                    if i == 1:
                        a = ws[coor].value
                    if i == 2:
                        b = ws[coor].value
                self.player = self.canvas.create_image(a, b, image = self.pSR)
                self.gravitated.append(self.player)
            if ws[r].value == "EP":
                rn = ws[r].row
                for i in range(1, 3):
                    char = chr(65 + i)
                    coor = char + str(rn)
                    if i == 1:
                        a = ws[coor].value
                    if i == 2:
                        b = ws[coor].value
                self.enemy = enemy.Enemies(a, b, self.canvas, "Pistol")
                self.enemyL.append(self.enemy)
            if ws[r].value == "EU":
                rn = ws[r].row
                for i in range(1, 3):
                    char = chr(65 + i)
                    coor = char + str(rn)
                    if i == 1:
                        a = ws[coor].value
                    if i == 2:
                        b = ws[coor].value
                self.enemy = enemy.Enemies(a, b, self.canvas, "Uzi")
                self.enemyL.append(self.enemy)
            if ws[r].value == "ESH":
                rn = ws[r].row
                for i in range(1, 3):
                    char = chr(65 + i)
                    coor = char + str(rn)
                    if i == 1:
                        a = ws[coor].value
                    if i == 2:
                        b = ws[coor].value
                self.enemy = enemy.Enemies(a, b, self.canvas, "Shotgun")
                self.enemyL.append(self.enemy)

    def wSelect(self, e):
        if self.gameOver:
            return
        n = int(e.char)
        if self.reloading:
            return 
        if n == 1 and "Pistol" in self.unlockedW:
            if self.direction == "l":
                self.canvas.itemconfig(self.player, image = self.pSL)
            if self.direction == "r":
                self.canvas.itemconfig(self.player, image = self.pSR)
            self.weapon = "Pistol"
            self.ammoC = 7
            self.reloadT = 20
            self.wCanvas.coords(self.check, 100, 100)
        if n == 2 and "Uzi" in self.unlockedW:
            if self.direction == "l":
                self.canvas.itemconfig(self.player, image = self.uSL)
            if self.direction == "r":
                self.canvas.itemconfig(self.player, image = self.uSR)
            self.weapon = "Uzi"
            self.reloadT = 35 
            self.ammoC = 32
            self.wCanvas.coords(self.check, 100, 230)
        elif "Uzi" not in self.unlockedW and n == 2:
            self.wLock("Uzi")
        if n == 3 and "Shotgun" in self.unlockedW:
            if self.direction == "l":
                self.canvas.itemconfig(self.player, image = self.shSL)
            if self.direction == "r":
                self.canvas.itemconfig(self.player, image = self.shSR)
            self.weapon = "Shotgun"
            self.reloadT = 80
            self.ammoC = 8
            self.wCanvas.coords(self.check, 100, 350)
        elif "Shotgun" not in self.unlockedW and n == 3:
            self.wLock("Shotgun")
        if n == 4 and "AR" in self.unlockedW:
            if self.direction == "l":
                self.canvas.itemconfig(self.player, image = self.arSL)
            if self.direction == "r":
                self.canvas.itemconfig(self.player, image = self.arSR)
            self.weapon = "AR"
            self.reloadT = 30
            self.ammoC = 30
            self.wCanvas.coords(self.check, 100, 460)
        elif "AR" not in self.unlockedW and n == 4:
            self.wLock("AR")
        if n == 5 and "Sniper" in self.unlockedW:
            if self.direction == "l":
                self.canvas.itemconfig(self.player, image = self.snSL)
            if self.direction == "r":
                self.canvas.itemconfig(self.player, image = self.snSR)
            self.weapon = "Sniper"
            self.reloadT = 50
            self.ammoC = 5
            self.wCanvas.coords(self.check, 100, 580)
        elif "Sniper" not in self.unlockedW and n == 5:
            self.wLock("Sniper")
            
    def wLock(self, w):
        if w == "Uzi":
            cost = 50
            lockPic = self.uL
            lockText = self.uT
        if w == "Shotgun":
            cost = 100
            lockPic = self.shL
            lockText = self.shT
        if w == "AR":
            cost = 200
            lockPic = self.arL
            lockText = self.arT
        if w == "Sniper":
            cost = 350
            lockPic = self.snL
            lockText = self.snT
        if w not in self.unlockedW and self.coins >= cost:
            self.wCanvas.delete(lockPic)
            self.wCanvas.delete(lockText)
            self.coins -= cost
            self.unlockedW.append(w)
            
    def shoot(self, e):
        if self.shootA > 0 and self.weapon == "Pistol" or self.shootA > 0 and self.weapon == "Shotgun" or self.shootA > 0 and self.weapon == "Sniper":
            return
        if self.reloading or self.gameOver or self.gHand or self.slashing:
            return
        if self.weapon == "Pistol":
            if self.pAmmo == 0:
                self.reload(self.weapon)
                return
            pxy = self.canvas.coords(self.player)
            x = pxy[0]
            y = pxy[1]
            self.bullet = bullets.Bullets(x, y, self.direction, self.canvas, self.weapon, "p")
            self.bulletL.append(self.bullet)
            self.shooting = True
            self.pAmmo -= 1
            self.shootA += 1
        if self.weapon == "Uzi":
            if self.uAmmo == 0:
                self.reload(self.weapon)
                return
            pxy = self.canvas.coords(self.player)
            x = pxy[0]
            y = pxy[1]
            self.bullet = bullets.Bullets(x, y, self.direction, self.canvas, self.weapon, "p")
            self.bulletL.append(self.bullet)
            self.shooting = True
            self.uAmmo -= 1
        if self.weapon == "Shotgun":
            if self.shAmmo == 0:
                self.reload(self.weapon)
                return
            pxy = self.canvas.coords(self.player)
            x = pxy[0]
            y = pxy[1]
            for i in range(8):
                self.bullet = bullets.Bullets(x, y, self.direction, self.canvas, self.weapon, "p")
                self.bulletL.append(self.bullet)
            self.shooting = True
            self.shAmmo -= 1
            self.shootA += 1
        if self.weapon == "AR":
            if self.arAmmo == 0:
                self.reload(self.weapon)
                return
            pxy = self.canvas.coords(self.player)
            x = pxy[0]
            y = pxy[1]
            self.bullet = bullets.Bullets(x, y, self.direction, self.canvas, self.weapon, "p")
            self.bulletL.append(self.bullet)
            self.shooting = True
            self.arAmmo -= 1
        if self.weapon == "Sniper":
            if self.snAmmo == 0:
                self.reload(self.weapon)
                return
            pxy = self.canvas.coords(self.player)
            x = pxy[0]
            y = pxy[1]
            self.bullet = bullets.Bullets(x, y, self.direction, self.canvas, self.weapon, "p")
            self.bulletL.append(self.bullet)
            self.shooting = True
            self.snAmmo -= 1
            self.shootA += 1
        self.totalScore += .01
            

    def shootR(self, e):
        if self.weapon == "Pistol" or self.weapon == "Shotgun" or self.weapon == "Sniper":
            self.shootA = 0

    def enemyShooting(self):
        pxy = self.canvas.coords(self.player)
        for i in self.enemyL:
            if i.ShootOrNah():
                PlaySound("SFX\Died.wav", SND_FILENAME and SND_ASYNC)
                return
            oxy = i.eLoc()
            if self.distCollision(pxy, oxy, 500):
                i.shoot()
                
    def swingK(self, e):
        if self.gameOver:
            return
        PlaySound("SFX\Swing.wav", SND_FILENAME and SND_ASYNC)
        pxy = self.canvas.coords(self.player)
        if self.direction == "r":
            self.canvas.itemconfig(self.player, image = self.kSR)
            self.slashing = True
        if self.direction == "l":
            self.canvas.itemconfig(self.player, image = self.kSL)
            self.slashing = True
        for i in self.enemyL:
            oxy = i.eLoc()
            if self.distCollision(pxy, oxy, 50):
                PlaySound("SFX\Stab.wav", SND_FILENAME and SND_ASYNC)
                i.shot(10)

    def reload(self, w):
        PlaySound("SFX\Empty.wav", SND_FILENAME and SND_ASYNC)
        if w == "Pistol":
            if self.reloadT > 0:
                self.reloading = True
                self.reloadT -= 1
                t = self.reloadT / 10
                self.canvas.itemconfig(self.rT, text = "Reloading \n" + str(t))
            elif self.reloadT == 0:
                self.reloading = False
                self.canvas.itemconfig(self.rT, text = "")
                self.pAmmo += 7
                self.reloadT += 20
        if w == "Uzi":
            if self.reloadT > 0:
                self.reloading = True
                self.reloadT -= 1
                t = self.reloadT / 10
                self.canvas.itemconfig(self.rT, text = "Reloading \n" + str(t))
            elif self.reloadT == 0:
                self.reloading = False
                self.canvas.itemconfig(self.rT, text = "")
                self.uAmmo += 32
                self.reloadT += 35
        if w == "Shotgun":
            if self.reloadT > 0:
                self.reloading = True
                self.reloadT -= 1
                t = self.reloadT / 10
                self.canvas.itemconfig(self.rT, text = "Reloading \n" + str(t))
            elif self.reloadT == 0:
                self.reloading = False
                self.canvas.itemconfig(self.rT, text = "")
                self.shAmmo += 8
                self.reloadT += 80
        if w == "AR":
            if self.reloadT > 0:
                self.reloading = True
                self.reloadT -= 1
                t = self.reloadT / 10
                self.canvas.itemconfig(self.rT, text = "Reloading \n" + str(t))
            elif self.reloadT == 0:
                self.reloading = False
                self.canvas.itemconfig(self.rT, text = "")
                self.arAmmo += 30
                self.reloadT += 30
        if w == "Sniper":
            if self.reloadT > 0:
                self.reloading = True
                self.reloadT -= 1
                t = self.reloadT / 10
                self.canvas.itemconfig(self.rT, text = "Reloading \n" + str(t))
            elif self.reloadT == 0:
                self.reloading = False
                self.canvas.itemconfig(self.rT, text = "")
                self.snAmmo += 5
                self.reloadT += 50
        self.totalScore += 1
        
    def jump(self, e):
        if self.gameOver:
            return
        if self.jumping or self.jAble == False:
            return
        self.jTimer = 15
        self.jumping = True

    def jumpR(self, e):
        self.jTimer = -1

    def a(self, e):
        if self.gameOver:
            return
        self.slashing = False
        self.direction = "l"
        if self.weapon == "Pistol":
            self.canvas.itemconfig(self.player, image = self.pSL)
        if self.weapon == "Uzi":
            self.canvas.itemconfig(self.player, image = self.uSL)
        if self.weapon == "Shotgun":
            self.canvas.itemconfig(self.player, image = self.shSL)
        if self.weapon == "AR":
            self.canvas.itemconfig(self.player, image = self.arSL)
        if self.weapon == "Sniper":
            self.canvas.itemconfig(self.player, image = self.snSL)
        self.x = -10

    def aR(self, e):
        self.x = 0

    def d(self, e):
        if self.gameOver:
            return
        self.slashing = False
        self.direction = "r"
        if self.weapon == "Pistol":
            self.canvas.itemconfig(self.player, image = self.pSR)
        if self.weapon == "Uzi":
            self.canvas.itemconfig(self.player, image = self.uSR)
        if self.weapon == "Shotgun":
            self.canvas.itemconfig(self.player, image = self.shSR)
        if self.weapon == "AR":
            self.canvas.itemconfig(self.player, image = self.arSR)
        if self.weapon == "Sniper":
            self.canvas.itemconfig(self.player, image = self.snSR)
        self.x = 10

    def dR(self, e):
        self.x = 0

    def distCollision(self, pxy, oxy, r):
        x = abs(pxy[0] - oxy[0])
        y = abs(pxy[1] - oxy[1])
        dist = math.sqrt(x * x + y * y)
        if dist < r:
            return True
        return False
    
    def aabbCollision(self, gxy, gH, gW, cxy, cH, cW, col):
        x = gxy[0]
        y = gxy[1]
        cx = cxy[0]
        cy = cxy[1]
        if col:
            cx = (cxy[0] + cxy[2]) / 2
            cy = (cxy[1] + cxy[3]) / 2        
            cH = cxy[3] - cxy[1]
            cW = cxy[2] - cxy[0]
        xCol = (math.fabs(x - cx) * 2) < (gH + cW)
        yCol = (math.fabs(y - cy) * 2) < (gW + cH)
        return (xCol and yCol)

    def collisions(self):
        for i in self.bulletL:
            xy = i.bLoc()
            for c in self.colliders:
                cxy = self.canvas.coords(c)
                if self.aabbCollision(xy, 3, 5, cxy, 0, 0, True):
                    i.hitFloors()
            for e in self.enemyL:
                exy = e.eLoc()
                if self.aabbCollision(xy, 3, 5, exy, 50, 66, False):
                    damage = i.getDamage()
                    e.shot(damage)
        for e in self.enemyL:
            enemyBL = e.enemyB()
            pxy = self.canvas.coords(self.player)
            for i in enemyBL:
                ebxy = i.bLoc()
                for c in self.colliders:
                    cxy = self.canvas.coords(c)
                    if self.aabbCollision(ebxy, 3, 5, cxy, 0, 0, True):
                        i.hitFloors()
                if self.aabbCollision(ebxy, 3, 5, pxy, 50, 66, False):
                    self.healthA -= 1
                    x = i.getDamage()
        for x in self.gravitated:
            gxy = self.canvas.coords(x)
            if gxy[1] > 750:
                self.healthA = 0
            for y in self.colliders:
                cxy = self.canvas.coords(y)
                if self.aabbCollision(gxy, 50, 66, cxy, 0, 0, True):
                    if gxy[1] < cxy[1]:
                        self.jAble = True
                        self.jumping = False
                        self.y = 0
                    elif gxy[1] > cxy[3]:
                        self.jTimer = -1
                    elif gxy[0] > cxy[0] or gxy[0] < cxy[2]:
                        self.x = 0
        for x in self.gravitated:
            gxy = self.canvas.coords(x)
            dxy = self.canvas.coords(self.door)
            if self.distCollision(gxy, dxy, 1020):
                self.playerMove = True
            if self.distCollision(gxy, dxy, 120):
                self.gameEnd("Win")

    def health(self):
        if self.healthA == 3:
            self.canvas.itemconfig(self.healthDisp, image = self.h3)
        elif self.healthA == 2:
            self.canvas.itemconfig(self.healthDisp, image = self.h2)
        elif self.healthA == 1:
            self.canvas.itemconfig(self.healthDisp, image = self.h1)
        else:
            self.canvas.itemconfig(self.healthDisp, image = self.h0)
            self.gameEnd("Loss")

    def heal(self, e):
        if self.gameOver:
            return
        if self.healthA < 3:
            self.ABA -= 1
            if self.ABA == 3:
                self.canvas.itemconfig(self.abDisp, image = self.AB3)
                PlaySound("SFX\Bandage.wav", SND_FILENAME and SND_ASYNC)
                self.healthA += 1
            elif self.ABA == 2:
                self.canvas.itemconfig(self.abDisp, image = self.AB2)
                PlaySound("SFX\Bandage.wav", SND_FILENAME and SND_ASYNC)
                self.healthA += 1
            elif self.ABA == 1:
                self.canvas.itemconfig(self.abDisp, image = self.AB1)
                PlaySound("SFX\Bandage.wav", SND_FILENAME and SND_ASYNC)
                self.healthA += 1
            elif self.ABA == 0:
                self.canvas.itemconfig(self.abDisp, image = self.ABX)
                PlaySound("SFX\Bandage.wav", SND_FILENAME and SND_ASYNC)
                self.healthA += 1
            self.totalScore += 1

    def gameEnd(self, wl):
        if self.gameOverR > 0:
            return
        self.gameOver = True
        if wl == "Loss":
            self.canvas.itemconfig(self.player, image = self.dS)
            self.canvas.create_image(640, 360, image = self.looseDisp)
            self.canvas.create_text(640, 350, text = "Score: " + str(int(self.totalScore)), font = ("Times", "45"))
            self.canvas.coords(self.coinDisp, 585, 470)
            self.canvas.tag_raise(self.coinDisp)
            self.gameOverR += 1
        if wl == "Win":
            self.win = True
            self.canvas.create_image(640, 400, image = self.cbg)
            self.Credits = self.canvas.create_image(640, 720, anchor = "n", image = self.cT)
            self.canvas.create_image(640,75, image = self.cfg)
            self.gameOverR += 1

    def ammoT(self):
        if self.weapon == "Pistol":
            return self.pAmmo
        if self.weapon == "Uzi":
            return self.uAmmo
        if self.weapon == "Shotgun":
            return self.shAmmo
        if self.weapon == "AR":
            return self.arAmmo
        if self.weapon == "Sniper":
            return self.snAmmo

    def endProgram(self):
        # Method called with stop buttion pressed
        self.stop = True

    def move(self):
        if self.gameOver:
            PlaySound(None, SND_PURGE)
            if self.win:
                print(self.canvas.coords(self.Credits))
                self.canvas.move(self.Credits, 0, -5)
            return
        if self.jTimer >= 0:
            self.y = -10
            self.jTimer -= 1
        else:
            self.y = 10
        if self.jTimer < 14:
            self.collisions()
        for i in self.enemyL:
            eBL = i.enemyB()
            for b in eBL:
                b.move()
                if b.removeB():
                    eBL.remove(b)
        for i in self.bulletL:
            i.move()
            if i.removeB():
                self.bulletL.remove(i)
        for i in self.enemyL:
            if i.removeE():
                self.coins += i.gMoney()
                self.totalScore += i.gMoney() * 1.5
                self.enemyL.remove(i)
        if self.playerMove == False:
            for i in self.enemyL:
                i.move(self.x)
            for i in self.colliders:
                self.canvas.move(i, self.x * -1, 0)
            for i in self.collidersI:
                self.canvas.move(i, self.x * -1, 0)
            self.canvas.move(self.door, self.x * -1, 0)
            self.canvas.move(self.player, 0, self.y)
        if self.playerMove:
            self.canvas.move(self.player, self.x, self.y)

    def main(self):
        # Main loop that runs the game
        while True:
            time.sleep(.1)
            if self.reloading == True:
                self.reload(self.weapon)
            t = str(self.ammoT())
            self.canvas.itemconfig(self.coinDisp, text = "Coins: " + str(self.coins))
            self.aDispT = self.weapon + "\n" + t + "/" + str(self.ammoC)
            self.canvas.itemconfig(self.ammoDisp, text = self.aDispT)
            self.enemyShooting()
            self.move()
            self.health()
            self.tk.update()
            self.tk.update_idletasks()
            if self.stop == True:
                self.tk.destroy()
                break

# Instantiate the class and run the main method
RNG = RunNGun()
RNG.main()
